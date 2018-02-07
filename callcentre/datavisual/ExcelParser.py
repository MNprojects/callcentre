import openpyxl
from .models import CallRecord, Document
from openpyxl import load_workbook


class ExcelParser():
       
    def read_excel(Document):
        wb2 = load_workbook('media/documents/Sheet1.xlsx' + os.pat Document)
        for row in wb2.active.iter_rows(range_string=wb2.active.calculate_dimension()):
            newRecord = CallRecord.objects.create(accountNumber = row[0].value, phoneNumber = row[1].value, 
            dialledNumber = row[2].value,
             destination = row[3].value,
              dateOfCall = row[4].value,
               timeOfCall = row[5].value,
              usageType = row[6].value,
               usageSubType = row[7].value,
              transmissionType = row[8].value, duration = row[9].value, uplinkVolume = row[10].value,
             downlinkVolume = row[11].value, totalVolume = row[12].value, cost = row[13].value, bundleType = row[14].value, roamedCategory = row[15].value, network =row[16].value, usageDirection = row[17].value, billSequenceNumber = row[18].value,
              invoiceNumber = row[19].value)
            newRecord.save()